import openpyxl

def readExcel(label,value):
    workbook = openpyxl.load_workbook("E:\\Data.xlsx")
    #print(workbook.sheetnames)
    sheet = workbook["TestDataSheet"]
    noofrows = sheet.max_row
    # print(noofrows)
    noofcols = sheet.max_column
    # print(noofcols)

    Label = []
    Header = []
    for i in range(1, noofrows + 1):
        Label.append(sheet.cell(i, 1).value)
        # print(Label[i-1])
        if Label[i - 1] == label:
            for j in range(1, noofcols + 1):
                Header.append(sheet.cell(1, j).value)
                # print(Header[j-1])
                if Header[j - 1] == value:
                    val = sheet.cell(i, j).value
                    break
            break
    return val


#print(readExcel("Microsoft","Region"))